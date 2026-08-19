"""
app/utils/excel_parser.py — Excel Ingestion Utility
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl

# Matches trailing price text embedded in "Item Name". The spec's example
# ("Chips Large - Rs.60") has a dash separator, but the real Bhopal export
# mostly omits it ("Phab ... choco truffle Rs. 60", "...Laddoo Rs.39") — the
# dash is therefore optional. Stripped to get the clean product name for
# matching against the product master.
_PRICE_SUFFIX_RE = re.compile(r"\s*-?\s*Rs\.?\s*\d+(\.\d+)?\s*$", re.IGNORECASE)

# Excel's date epoch (serial day 0) — needed because some exports mix real
# datetime cells with raw Excel serial-number cells in the same DateVal
# column (observed in the real Bhopal July export: ~9% of rows are serials).
_EXCEL_EPOCH = datetime(1899, 12, 30)


def clean_item_name(raw_name: str) -> str:
    """Strip trailing '- Rs.NN' price text from an Item Name export value."""
    if not raw_name:
        return ""
    return _PRICE_SUFFIX_RE.sub("", raw_name).strip()


def _excel_serial_to_date(value: float) -> date:
    return (_EXCEL_EPOCH + __import__("datetime").timedelta(days=value)).date()


def _coerce_date(value: Any) -> date | None:
    """Handle DateVal cells that may be a datetime, a date, or a raw Excel
    serial number (int/float) — all three appear in real exports."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return _excel_serial_to_date(float(value))
        except (OverflowError, ValueError):
            return None
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _coerce_time(value: Any) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return None


def _to_decimal(value: Any, default: str = "0") -> Decimal:
    if value is None or value == "":
        return Decimal(default)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _to_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def parse_vend_transactions_excel(file_bytes: bytes, sheet_name: str = "July") -> dict[str, Any]:
    """
    Parses the raw vending-machine sales export (one row per order/vend
    attempt) into normalized dicts ready for upsert into `machines`,
    `products`, and `vend_transactions`.

    Expected columns (Vendiman spec section 3) — matched case-insensitively,
    with both space and underscore variants accepted:
    OrderId, OrderNo, OperationLocationName, Client Id, Client Name,
    Client Location Name, TelemetrySerialNo, TerminalSerialNo, MachineId,
    MachineName, MachineUID, Date, Time, Amount, PaidAmount, Order Quantity,
    Success Quantity, Failed Quantity, Remarks, Status, PaymentStatus,
    RefundStatus, RefundedAmount, Payment/refund Remarks, PaymentMode,
    Prepaid Card Number, Prepaid Card Uid, Item, ReferenceNo,
    UTR/BANK_TRANSFER_NO, BinNumber, OPCode, C_Prd_MRP, Source, Device Model,
    Device OS, Device OS Version, Hour, DateVal, Weekday, WeekType,
    Item Name, Net Revenue, MachineFail, Lost Value (MRP x FailedQty)

    Returns:
        {
            "rows": [ { ...cleaned row... }, ... ],   # one per valid source row
            "machines": { machine_uid: {...machine fields...} },  # deduped
            "skipped": int,                             # rows with no machine_uid
            "total_rows": int,
        }

    Each cleaned row carries both `product_name_clean` (price-stripped, used
    to match/create the product master) and `raw_item_name` (untouched, kept
    on the transaction row for audit/reconciliation of match failures).
    """
    # read_only=True is important here: the source workbook typically ships
    # with several other heavy dashboard/pivot sheets alongside the raw data
    # sheet, and loading the whole workbook in normal mode can use several
    # hundred MB of RAM for a 100k-row file. Read-only mode streams rows
    # instead of materializing the whole sheet grid.
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    sheet = wb[sheet_name]

    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    machines: dict[str, dict[str, Any]] = {}
    skipped = 0
    total = 0

    def col(row_dict: dict[str, Any], *names: str) -> Any:
        for n in names:
            key = n.strip().lower()
            if key in row_dict:
                return row_dict[key]
        return None

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(c).strip().lower() if c is not None else f"unnamed_{i}" for i, c in enumerate(row)]
            continue
        if not any(row):
            continue

        raw = dict(zip(headers, row))
        total += 1

        machine_uid = _to_str(col(raw, "machineuid", "machine uid"))
        if not machine_uid:
            skipped += 1
            continue

        # Register/refresh the machine this row belongs to
        if machine_uid not in machines:
            machines[machine_uid] = {
                "machine_uid": machine_uid,
                "source_machine_id": _to_str(col(raw, "machineid", "machine id")),
                "name": _to_str(col(raw, "machinename", "machine name")) or machine_uid,
                "telemetry_serial_no": _to_str(col(raw, "telemetryserialno", "telemetry serial no")),
                "terminal_serial_no": _to_str(col(raw, "terminalserialno", "terminal serial no")),
                "operation_location_name": _to_str(col(raw, "operationlocationname", "operation location name")),
                "client_source_id": _to_str(col(raw, "client id", "client_id")),
                "client_name": _to_str(col(raw, "client name", "client_name")),
                "client_location_name": _to_str(col(raw, "client location name", "client_location_name")),
                "device_model": _to_str(col(raw, "device model", "device_model")),
                "device_os": _to_str(col(raw, "device os", "device_os")),
                "device_os_version": _to_str(col(raw, "device os version", "device_os_version")),
            }

        raw_item_name = _to_str(col(raw, "item name", "item_name")) or _to_str(col(raw, "item")) or ""
        product_name_clean = clean_item_name(raw_item_name)

        machine_fail_val = col(raw, "machinefail", "machine fail")
        machine_fail = bool(machine_fail_val) and str(machine_fail_val).strip().lower() not in ("0", "false", "no", "")

        rows.append({
            "source_order_id": _to_str(col(raw, "orderid", "order id")),
            "order_no": _to_str(col(raw, "orderno", "order no")),
            "machine_uid": machine_uid,
            "raw_item_name": raw_item_name,
            "product_name_clean": product_name_clean,
            "bin_number": _to_str(col(raw, "binnumber", "bin number")),
            "op_code": _to_str(col(raw, "opcode", "op code")),

            "txn_date": _coerce_date(col(raw, "dateval")) or _coerce_date(col(raw, "date")),
            "txn_time": _coerce_time(col(raw, "time")),
            "hour_of_day": _to_int(col(raw, "hour"), default=None) if col(raw, "hour") not in (None, "") else None,
            "weekday": _to_str(col(raw, "weekday")),
            "week_type": _to_str(col(raw, "weektype", "week type")),

            "order_quantity": _to_int(col(raw, "order quantity", "order_quantity")),
            "success_quantity": _to_int(col(raw, "success quantity", "success_quantity")),
            "failed_quantity": _to_int(col(raw, "failed quantity", "failed_quantity")),
            "machine_fail": machine_fail,
            "lost_value": _to_decimal(col(raw, "lost value (mrp x failedqty)", "lost value")),

            "amount": _to_decimal(col(raw, "amount")),
            "paid_amount": _to_decimal(col(raw, "paidamount", "paid amount")),
            "net_revenue": _to_decimal(col(raw, "net revenue", "net_revenue")),
            "mrp_at_sale": _to_decimal(col(raw, "c_prd_mrp")),

            "status": _to_str(col(raw, "status")),
            "payment_status": _to_str(col(raw, "paymentstatus", "payment status")),
            "refund_status": _to_str(col(raw, "refundstatus", "refund status")),
            "refunded_amount": _to_decimal(col(raw, "refundedamount", "refunded amount")),
            "remarks": _to_str(col(raw, "remarks")),
            "payment_refund_remarks": _to_str(col(raw, "payment/refund remarks")),

            "payment_mode": _to_str(col(raw, "paymentmode", "payment mode")),
            "prepaid_card_number": _to_str(col(raw, "prepaid card number")),
            "prepaid_card_uid": _to_str(col(raw, "prepaid card uid")),
            "reference_no": _to_str(col(raw, "referenceno", "reference no")),
            "utr_bank_transfer_no": _to_str(col(raw, "utr/bank_transfer_no")),

            "source": _to_str(col(raw, "source")),
            "device_model": _to_str(col(raw, "device model")),
            "device_os": _to_str(col(raw, "device os")),
            "device_os_version": _to_str(col(raw, "device os version")),

            "mrp_for_product": _to_decimal(col(raw, "c_prd_mrp")),
        })

    return {
        "rows": rows,
        "machines": machines,
        "skipped": skipped,
        "total_rows": total,
    }


def parse_product_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Parses an uploaded Excel sheet of products.
    Expected columns (case-insensitive):
    SKU, Name, Category, MRP, Cost Price, Selling Price, GST Rate, Unit, Reorder Point, Reorder Quantity
    """
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    headers = []
    rows = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            continue

        if not any(row):
            continue

        row_dict = dict(zip(headers, row))
        sku = str(row_dict.get("sku", "") or "").strip().upper()
        name = str(row_dict.get("name", "") or "").strip()

        if not sku or not name:
            continue

        try:
            mrp = Decimal(str(row_dict.get("mrp", 0) or 0))
            cost_price = Decimal(str(row_dict.get("cost price", row_dict.get("cost_price", 0)) or 0))
            selling_price = Decimal(str(row_dict.get("selling price", row_dict.get("selling_price", mrp)) or mrp))
            gst = str(row_dict.get("gst rate", row_dict.get("gst_rate", "18")) or "18").replace("%", "").strip()
            reorder_point = int(row_dict.get("reorder point", row_dict.get("reorder_point", 10)) or 10)
            reorder_qty = int(row_dict.get("reorder quantity", row_dict.get("reorder_quantity", 50)) or 50)
            brand = str(row_dict.get("brand", "") or "").strip() or None
            unit = str(row_dict.get("unit", "piece") or "piece").strip()

            rows.append({
                "sku": sku,
                "name": name,
                "brand": brand,
                "unit": unit,
                "mrp": mrp,
                "cost_price": cost_price,
                "selling_price": selling_price,
                "gst_rate": gst,
                "reorder_point": reorder_point,
                "reorder_quantity": reorder_qty,
            })
        except Exception:
            continue

    return rows
